import sys
import os
import webbrowser
import threading
import time
import json
from flask import Flask, render_template, request, jsonify, redirect, url_for, session

# Añadir el directorio raíz al path para poder importar las clases de models
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from models.usuarios.Clase_Estudiante import Estudiante
from models.usuarios.Clase_Docente import Docente
from models.usuarios.Clase_Coordinador import Coordinador
from models.academico.Clase_Materia import Materia
from models.academico.Clase_Seccion import Seccion
from models.academico.Clase_Periodo import Periodo
from models.academico.Clase_Horario import Horario
from models.patrones_diseno.bridge.AulaVirtualBridge import AulaVirtual, AulaClaseSincrona, AulaExamen, ServicioTeams, ServicioZoom
from models.enums.Estado_Aprobacion import EstadoDeAprobacionMateria, EstadoDeAprobacionNivelacion
from models.institucion.Clase_Carrera import Carrera
from models.patrones_diseno.strategy.ReporteStrategy import Reporte
from models.gestion.Clase_NotaMateria import NotaMateria
from models.patrones_diseno.builders.SeccionBuilder import SeccionBuilder

app = Flask(__name__)
app.secret_key = 'super_secret_key_for_webview_app'

# Ruta absoluta al archivo JSON
DB_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), 'database', 'db.json'))

# =========================================================================
# VARIABLES GLOBALES DE INSTANCIAS POO
# =========================================================================
periodo_actual = None
materias = {}
carrera_software = None
carreras = {}
periodos = {}
estudiantes = {}
docentes = {}
coordinadores = {}
secciones = {}
solicitudes_retiro = []
reportes_generados = []

# =========================================================================
# CARGA Y GUARDADO DE LA BASE DE DATOS JSON A OBJETOS POO
# =========================================================================
def init_career_db(db_dir, id_carrera, nombre_carrera):
    career_dir = os.path.join(db_dir, id_carrera)
    if not os.path.exists(career_dir):
        os.makedirs(career_dir)
        print(f"[Sistema] Creando carpeta de base de datos para la carrera {nombre_carrera} ({id_carrera})...")
        
    for sub in ['students', 'mallas_curriculares']:
        s_path = os.path.join(career_dir, sub)
        if not os.path.exists(s_path):
            os.makedirs(s_path)
            
    defaults = {
        'periodos.json': [{
            "nombre_periodo": f"Nivelacion {nombre_carrera}",
            "fecha_inicio": "2026-01-01",
            "fecha_final": "2026-06-01",
            "estado_periodo": "Planificación"
        }],
        'docentes.json': [],
        'secciones.json': [],
        'retiros.json': [],
        'notas.json': []
    }
    
    for filename, content in defaults.items():
        file_path = os.path.join(career_dir, filename)
        if not os.path.exists(file_path):
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(content, f, indent=4, ensure_ascii=False)


def load_db():
    global periodo_actual, materias, carrera_software, carreras, periodos, estudiantes, docentes, coordinadores, secciones, solicitudes_retiro, reportes_generados
    
    db_dir = os.path.join(os.path.dirname(__file__), 'database')
    
    def load_json_root(filename):
        path = os.path.join(db_dir, filename)
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []

    # 1. Carreras (root)
    c_data_list = load_json_root('carreras.json')
    carreras = {}
    if c_data_list:
        for c_data in c_data_list:
            c_obj = Carrera(
                id_carrera=c_data["id_carrera"],
                nombre_carrera=c_data["nombre_carrera"],
                capacidad_estudiantil=c_data["capacidad_estudiantil"]
            )
            carreras[c_obj.id_carrera] = c_obj
        carrera_software = list(carreras.values())[0] if carreras else Carrera("C-001", "Software", 500)
    else:
        carrera_software = Carrera("C-001", "Software", 500)
        carreras[carrera_software.id_carrera] = carrera_software

    # 2. Inicialización/migración
    # Si existe periodos.json en raíz, realizamos la migración a C-001 (Software)
    root_periodos_file = os.path.join(db_dir, 'periodos.json')
    if os.path.exists(root_periodos_file):
        print("[Sistema] Detectados datos en el directorio raíz. Ejecutando migración a C-001...")
        target_dir = os.path.join(db_dir, 'C-001')
        os.makedirs(target_dir, exist_ok=True)
        import shutil
        for f in ['periodos.json', 'docentes.json', 'secciones.json', 'retiros.json', 'notas.json']:
            src = os.path.join(db_dir, f)
            if os.path.exists(src):
                shutil.copy(src, os.path.join(target_dir, f))
                os.remove(src)
        for d in ['students', 'mallas_curriculares']:
            src_dir = os.path.join(db_dir, d)
            dst_dir = os.path.join(target_dir, d)
            if os.path.exists(src_dir):
                if os.path.exists(dst_dir):
                    shutil.rmtree(dst_dir)
                shutil.copytree(src_dir, dst_dir)
                shutil.rmtree(src_dir)
        # Limpiar otros archivos viejos que ya se migraron
        for f in ['notas.json', 'retiros.json', 'secciones.json', 'docentes.json']:
            src = os.path.join(db_dir, f)
            if os.path.exists(src):
                os.remove(src)
                
    # Inicializar directorios de carrera para todas las carreras registradas
    for c_id, c_obj in carreras.items():
        init_career_db(db_dir, c_id, c_obj.nombre_carrera)

    # 3. Coordinadores (loaded from root)
    coordinadores = {}
    for c in load_json_root('coordinadores.json'):
        coord = Coordinador(
            cedula=c["cedula"],
            nombres=c["nombres"],
            apellidos=c["apellidos"],
            correo=c["correo"],
            contrasenia=c["contrasenia"],
            id_coordinador=c["id_coordinador"],
            fecha_asignacion_cargo=c["fecha_asignacion_cargo"]
        )
        id_carrera = c.get("id_carrera")
        if id_carrera and id_carrera in carreras:
            coord.asociar_carrera(carreras[id_carrera])
        else:
            coord.asociar_carrera(carrera_software)
        coordinadores[coord._correo] = coord

    # Limpiar variables globales planas
    periodos = {}
    materias = {}
    docentes = {}
    estudiantes = {}
    secciones = {}
    solicitudes_retiro = []

    # 4. Cargar datos específicos por carrera
    for c_id in carreras.keys():
        career_dir = os.path.join(db_dir, c_id)
        
        def load_json_career(filename):
            path = os.path.join(career_dir, filename)
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return []

        # Periodo de esta carrera
        p_data_list = load_json_career('periodos.json')
        if p_data_list:
            p_data = p_data_list[0]
            p_obj = Periodo(
                nombre_periodo=p_data["nombre_periodo"],
                fecha_inicio=p_data["fecha_inicio"],
                fecha_final=p_data["fecha_final"]
            )
            from models.enums.Estado_Periodo import EstadoPeriodo
            for state_enum in EstadoPeriodo:
                if state_enum.value == p_data.get("estado_periodo"):
                    p_obj._estado_periodo = state_enum
                    break
            periodos[c_id] = p_obj
        else:
            periodos[c_id] = Periodo("Periodo", "2024", "2024")

        # Materias de esta carrera
        mallas_dir = os.path.join(career_dir, 'mallas_curriculares')
        import glob
        if os.path.exists(mallas_dir):
            for m_file in glob.glob(os.path.join(mallas_dir, '*.json')):
                with open(m_file, 'r', encoding='utf-8') as f:
                    m_list = json.load(f)
                    for m in m_list:
                        m_obj = Materia(
                            id_materia=m["id_materia"],
                            nombre_materia=m["nombre_materia"],
                            nota_minima=m["nota_minima"],
                            asistencia_minima=m["asistencia_minima"]
                        )
                        m_obj.carrera_id = c_id
                        materias[m_obj.id_materia] = m_obj

        # Docentes de esta carrera
        for d in load_json_career('docentes.json'):
            doc = Docente(
                cedula=d["cedula"],
                nombres=d["nombres"],
                apellidos=d["apellidos"],
                correo=d["correo"],
                contrasenia=d["contrasenia"],
                especialidad=d.get("especialidad", "")
            )
            doc.carrera_id = c_id
            docentes[doc._correo] = doc

        # Estudiantes de esta carrera
        students_dir = os.path.join(career_dir, 'students')
        if os.path.exists(students_dir):
            est_files = glob.glob(os.path.join(students_dir, '*.json'))
            est_list = []
            for file_path in est_files:
                filename = os.path.basename(file_path)
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    for e in data:
                        e['_archivo_origen'] = filename
                        est_list.append(e)
            
            for e in est_list:
                est = Estudiante(
                    cedula=e["cedula"],
                    nombres=e["nombres"],
                    apellidos=e["apellidos"],
                    correo=e["correo"],
                    contrasenia=e["contrasenia"],
                    id_estudiante=e["id_estudiante"],
                    nombre_periodo=e["nombre_periodo"],
                    tipo_matricula=e["tipo_matricula"]
                )
                est.esta_activo = e.get("esta_activo", 1)
                est.notificaciones = e.get("notificaciones", [])
                est._archivo_origen = e.get("_archivo_origen", "estudiantes.json")
                est.carrera_id = c_id
                estudiantes[est._correo] = est

        # Secciones de esta carrera
        for s in load_json_career('secciones.json'):
            m_obj = materias.get(s["materia_id"])
            coord_correo = s.get("coordinador_correo")
            coord_obj = coordinadores.get(coord_correo) if coord_correo else next((co for co in coordinadores.values() if co.carrera and co.carrera.id_carrera == c_id), None)
            if not coord_obj and coordinadores:
                coord_obj = list(coordinadores.values())[0]

            sec = (SeccionBuilder()
                   .con_id_seccion(s["id_seccion"])
                   .con_capacidad_estudiantil(s["capacidad_estudiantil"])
                   .con_materia(m_obj)
                   .con_coordinador(coord_obj)
                   .build())
            sec.carrera_id = c_id
                   
            if s.get("aula_virtual"):
                av = s["aula_virtual"]
                sec.aula_virtual = AulaVirtual(
                    capacidad_maxima=av["capacidad_maxima"],
                    enlace_plataforma=av["enlace_plataforma"],
                    tipo_plataforma=av["tipo_plataforma"]
                )
                
            for doc_correo in s.get("docentes_correos", []):
                doc_obj = docentes.get(doc_correo)
                if doc_obj:
                    sec.asignar_docente(doc_obj)
                    
            for h in s.get("horarios", []):
                hor = Horario(
                    turno=h["turno"],
                    hora_inicio=h["hora_inicio"],
                    hora_fin=h["hora_fin"],
                    modalidad="PRESENCIAL" if h["modalidad"] == "PRESENCIAL" else "VIRTUAL",
                    dias=h["dias"]
                )
                sec.agregar_horario(hor)
                
            for est_id in s.get("estudiantes_ids", []):
                est_obj = next((e for e in estudiantes.values() if e._id_estudiante == est_id and getattr(e, 'carrera_id', None) == c_id), None)
                if est_obj:
                    sec.actualizar_estudiantes_inscritos(est_obj)
                    if sec not in est_obj.secciones_asociadas:
                        est_obj.secciones_asociadas.append(sec)
                        
            secciones[sec.id_seccion] = sec

        # Notas de esta carrera
        for n in load_json_career('notas.json'):
            est_obj = next((e for e in estudiantes.values() if e._id_estudiante == n["estudiante_id"] and getattr(e, 'carrera_id', None) == c_id), None)
            m_obj = materias.get(n["materia_id"])
            if est_obj and m_obj:
                est_obj.historial.crear_nota_materia(
                    materia=m_obj,
                    periodo=periodos[c_id],
                    parcial1=n["parcial1"],
                    parcial2=n["parcial2"],
                    asistencia=n["asistencia"]
                )
                nota_obj = est_obj.historial.lista_nota_materia[-1]
                nota_obj.periodo_cerrado = n.get("periodo_cerrado", False)

        # Auto-sanación de notas para esta carrera
        for est_obj in estudiantes.values():
            if getattr(est_obj, 'carrera_id', None) == c_id:
                for sec in est_obj.secciones_asociadas:
                    if sec.materia:
                        nota_obj = next((n for n in est_obj.historial.lista_nota_materia if n.materia.id_materia == sec.materia.id_materia), None)
                        if not nota_obj:
                            est_obj.historial.crear_nota_materia(
                                materia=sec.materia,
                                periodo=periodos[c_id],
                                parcial1=0.0,
                                parcial2=0.0,
                                asistencia=0
                            )

        # Retiros de esta carrera
        for sol in load_json_career('retiros.json'):
            if sol.get("reporte") and isinstance(sol["reporte"], dict):
                rep_data = sol["reporte"]
                sol["reporte"] = Reporte(
                    tipo_de_reporte=rep_data.get("tipo_de_reporte"),
                    formato_documento=rep_data.get("formato_documento"),
                    emisor=rep_data.get("emisor"),
                    contenido=rep_data.get("contenido")
                )
            sol["id_carrera"] = c_id
            solicitudes_retiro.append(sol)

    if periodos:
        periodo_actual = list(periodos.values())[0]


def save_db():
    db_dir = os.path.join(os.path.dirname(__file__), 'database')
    
    def save_json_root(filename, data):
        with open(os.path.join(db_dir, filename), 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

    # 1. Carreras (root)
    c_data = []
    for c_obj in carreras.values():
        c_data.append({
            "id_carrera": c_obj.id_carrera,
            "nombre_carrera": c_obj.nombre_carrera,
            "capacidad_estudiantil": c_obj.capacidad_estudiantil
        })
    save_json_root('carreras.json', c_data)

    # 2. Coordinadores (root)
    coord_list = []
    for c in coordinadores.values():
        coord_list.append({
            "cedula": c.cedula,
            "nombres": c.nombres,
            "apellidos": c.apellidos,
            "correo": c._correo,
            "contrasenia": c.contrasenia,
            "id_coordinador": c.id_coordinador,
            "fecha_asignacion_cargo": c.fecha_asignacion_cargo,
            "id_carrera": c.carrera.id_carrera if c.carrera else None
        })
    save_json_root('coordinadores.json', coord_list)

    # 3. Guardar datos por carrera
    for c_id in carreras.keys():
        career_dir = os.path.join(db_dir, c_id)
        os.makedirs(career_dir, exist_ok=True)
        
        def save_json_career(filename, data):
            with open(os.path.join(career_dir, filename), 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)

        # Periodo
        p_obj = periodos.get(c_id)
        if p_obj:
            p_data = [{
                "nombre_periodo": p_obj.nombre_periodo,
                "fecha_inicio": p_obj.fecha_inicio,
                "fecha_final": p_obj.fecha_final,
                "estado_periodo": p_obj._estado_periodo.value if hasattr(p_obj, '_estado_periodo') else "En curso"
            }]
            save_json_career('periodos.json', p_data)

        # Docentes de esta carrera
        doc_list = []
        for d in docentes.values():
            if getattr(d, 'carrera_id', None) == c_id:
                doc_list.append({
                    "cedula": d.cedula,
                    "nombres": d.nombres,
                    "apellidos": d.apellidos,
                    "correo": d._correo,
                    "contrasenia": d.contrasenia,
                    "especialidad": getattr(d, 'especialidad', '')
                })
        save_json_career('docentes.json', doc_list)

        # Estudiantes de esta carrera
        from collections import defaultdict
        est_grouped = defaultdict(list)
        for e in estudiantes.values():
            if getattr(e, 'carrera_id', None) == c_id:
                archivo_destino = getattr(e, '_archivo_origen', 'estudiantes.json')
                est_grouped[archivo_destino].append({
                    "cedula": e.cedula,
                    "nombres": e.nombres,
                    "apellidos": e.apellidos,
                    "correo": e._correo,
                    "contrasenia": e.contrasenia,
                    "id_estudiante": e._id_estudiante,
                    "nombre_periodo": e.nombre_periodo,
                    "tipo_matricula": e._tipo_matricula,
                    "esta_activo": getattr(e, "esta_activo", 1),
                    "notificaciones": getattr(e, "notificaciones", [])
                })
        students_dir = os.path.join(career_dir, 'students')
        os.makedirs(students_dir, exist_ok=True)
        for filename, est_list in est_grouped.items():
            with open(os.path.join(students_dir, filename), 'w', encoding='utf-8') as f:
                json.dump(est_list, f, indent=4, ensure_ascii=False)

        # Secciones de esta carrera
        sec_list = []
        for s in secciones.values():
            if getattr(s, 'carrera_id', None) == c_id:
                doc_emails = [d._correo for d in s.docentes]
                est_ids = [e._id_estudiante for e in s.estudiantes_inscritos]
                horarios_list = []
                for h in s.lista_horarios:
                    horarios_list.append({
                        "turno": h.turno,
                        "hora_inicio": h.hora_inicio,
                        "hora_fin": h.hora_fin,
                        "modalidad": h._modalidad,
                        "dias": h.dias
                    })
                av_dict = None
                if s.aula_virtual:
                    av_dict = {
                        "capacidad_maxima": s.aula_virtual.capacidad_maxima,
                        "enlace_plataforma": s.aula_virtual._enlace_plataforma,
                        "tipo_plataforma": s.aula_virtual._tipo_plataforma
                    }
                sec_list.append({
                    "id_seccion": s.id_seccion,
                    "capacidad_estudiantil": s.capacidad_estudiantil,
                    "materia_id": s.materia.id_materia if s.materia else "",
                    "coordinador_correo": s.coordinador._correo if s.coordinador else None,
                    "docentes_correos": doc_emails,
                    "estudiantes_ids": est_ids,
                    "horarios": horarios_list,
                    "aula_virtual": av_dict
                })
        save_json_career('secciones.json', sec_list)

        # Notas de esta carrera
        notas_list = []
        for est in estudiantes.values():
            if getattr(est, 'carrera_id', None) == c_id:
                for n in est.historial.lista_nota_materia:
                    notas_list.append({
                        "estudiante_id": est._id_estudiante,
                        "materia_id": n.materia.id_materia,
                        "parcial1": n.parcial1,
                        "parcial2": n.parcial2,
                        "asistencia": n.asistencia,
                        "periodo_cerrado": getattr(n, 'periodo_cerrado', False)
                    })
        save_json_career('notas.json', notas_list)

        # Retiros de esta carrera
        retiros_list = []
        for sol in solicitudes_retiro:
            if sol.get('id_carrera') == c_id:
                sol_copy = dict(sol)
                rep_dict = None
                if sol_copy.get("reporte"):
                    if isinstance(sol_copy["reporte"], Reporte):
                        rep_dict = {
                            "tipo_de_reporte": sol_copy["reporte"].tipo_de_reporte,
                            "formato_documento": sol_copy["reporte"].formato_documento,
                            "emisor": sol_copy["reporte"].emisor,
                            "contenido": sol_copy["reporte"].contenido
                        }
                    else:
                        rep_dict = sol_copy["reporte"]
                sol_copy["reporte"] = rep_dict
                sol_copy.pop('id_carrera', None)
                retiros_list.append(sol_copy)
        save_json_career('retiros.json', retiros_list)


def obtener_usuario_por_correo(correo):
    # Buscar en Coordinadores
    user = coordinadores.get(correo)
    if user:
        return user, 'coordinador'
        
    # Buscar en Docentes
    user = docentes.get(correo)
    if user:
        return user, 'docente'
        
    # Buscar en Estudiantes
    user = estudiantes.get(correo)
    if user:
        return user, 'estudiante'
        
    return None, None

def get_current_period():
    c_id = session.get('carrera_id')
    if c_id and c_id in periodos:
        return periodos[c_id]
    return periodo_actual

def verificar_contrasenia(usuario, contrasenia_ingresada):
    if not usuario:
        return False
    return usuario.contrasenia == contrasenia_ingresada

# Nota: La propiedad 'esta_aprobado' se encuentra definida internamente en NotaMateria
# usando el patrón Observer y la verificación del estado del periodo. No se inyecta dinámicamente.


# =========================================================================
# RUTAS DE FLASK
# =========================================================================

@app.route('/')
def index():
    if 'usuario' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    correo = request.form.get('correo', '').strip()
    contrasenia = request.form.get('contrasenia', '').strip()

    usuario, rol = obtener_usuario_por_correo(correo)
    if usuario and verificar_contrasenia(usuario, contrasenia):
        if rol == 'estudiante' and getattr(usuario, 'esta_activo', 1) == 0:
            return jsonify({"status": "error", "message": "No registrado"})
            
        session['usuario'] = correo
        session['rol'] = rol
        session['nombre'] = usuario.obtener_nombre_completo()
        
        if rol == 'estudiante':
            session['id'] = usuario._id_estudiante
            session['carrera_id'] = getattr(usuario, 'carrera_id', None)
        elif rol == 'coordinador':
            session['id'] = usuario.id_coordinador
            session['carrera_id'] = usuario.carrera.id_carrera if usuario.carrera else None
        else:
            session['id'] = usuario.cedula
            session['carrera_id'] = getattr(usuario, 'carrera_id', None)
            
        return jsonify({"status": "success", "redirect": url_for('dashboard')})
    else:
        return jsonify({"status": "error", "message": "Credenciales inválidas. Por favor verifique el correo y la contraseña."})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/dashboard')
def dashboard():
    if 'usuario' not in session:
        return redirect(url_for('index'))
    
    rol = session['rol']
    if rol == 'estudiante':
        return redirect(url_for('dashboard_estudiante'))
    elif rol == 'docente':
        return redirect(url_for('dashboard_docente'))
    elif rol == 'coordinador':
        return redirect(url_for('dashboard_coordinador'))
    return redirect(url_for('logout'))

# -------------------------------------------------------------------------
# RUTA ESTUDIANTE
# -------------------------------------------------------------------------
@app.route('/student')
def dashboard_estudiante():
    if 'usuario' not in session or session['rol'] != 'estudiante':
        return redirect(url_for('index'))
    
    est = estudiantes.get(session['usuario'])
    if not est:
        return redirect(url_for('logout'))
        
    carrera_id = getattr(est, 'carrera_id', None)
    p_actual = periodos.get(carrera_id) or periodo_actual
    
    # Calificaciones
    notas_info = []
    # Asegurar que cada sección asociada tenga su objeto NotaMateria
    for sec in est.secciones_asociadas:
        if sec.materia:
            nota_obj = next((n for n in est.historial.lista_nota_materia if n.materia.id_materia == sec.materia.id_materia), None)
            if not nota_obj:
                est.historial.crear_nota_materia(
                    materia=sec.materia,
                    periodo=p_actual,
                    parcial1=0.0,
                    parcial2=0.0,
                    asistencia=0
                )

    for nota in est.historial.lista_nota_materia:
        estado_m = nota.esta_aprobado
        notas_info.append({
            "codigo": nota.materia.id_materia,
            "materia": nota.materia.nombre_materia,
            "parcial1": nota.parcial1,
            "parcial2": nota.parcial2,
            "nota_final": nota.nota_final,
            "asistencia": nota.asistencia,
            "estado": estado_m.value,
            "badge_class": "badge-success" if estado_m == EstadoDeAprobacionMateria.MATERIA_APROBADA else ("badge-danger" if estado_m == EstadoDeAprobacionMateria.MATERIA_REPROBADA else "badge-warning")
        })

    veredicto = est.esta_aprobado
    veredicto_badge = "badge-success" if veredicto == EstadoDeAprobacionNivelacion.APROBADO else ("badge-danger" if veredicto == EstadoDeAprobacionNivelacion.REPROBADO else "badge-warning")

    # Horarios
    horarios_info = est.ver_horario()

    mis_reportes = [r for r in reportes_generados if r.emisor == est.obtener_nombre_completo()]

    return render_template(
        'dashboard_estudiante.html',
        estudiante=est,
        notas=notas_info,
        promedio=est.historial.promedio_general,
        veredicto=veredicto.value,
        veredicto_badge=veredicto_badge,
        horarios=horarios_info,
        reportes=mis_reportes,
        periodo=p_actual.nombre_periodo,
        estado_periodo=p_actual.estado_periodo
    )

@app.route('/student/request_certificate', methods=['POST'])
def student_request_certificate():
    if 'usuario' not in session or session['rol'] != 'estudiante':
        return jsonify({"status": "error", "message": "No autorizado"})
    
    est = estudiantes[session['usuario']]
    formato = request.form.get('formato', 'PDF')
    
    reporte = est.solicitar_certificado(formato_documento=formato)
    reportes_generados.append(reporte)
    
    save_db() # Guardar base de datos
    return jsonify({
        "status": "success", 
        "message": "Certificado solicitado con éxito", 
        "report_content": reporte.imprimir_reporte()
    })

@app.route('/student/request_withdrawal', methods=['POST'])
def student_request_withdrawal():
    if 'usuario' not in session or session['rol'] != 'estudiante':
        return jsonify({"status": "error", "message": "No autorizado"})
    
    est = estudiantes[session['usuario']]
    motivo = request.form.get('motivo', '').strip()
    formato = request.form.get('formato', 'Consola')
    
    if not motivo:
        return jsonify({"status": "error", "message": "Debe especificar un motivo para el retiro."})
    
    reporte = est.solicitar_retiro(motivo=motivo, formato_documento=formato)
    
    solicitudes_retiro.append({
        "id": len(solicitudes_retiro) + 1,
        "estudiante_id": est._id_estudiante,
        "nombre": est.obtener_nombre_completo(),
        "correo": est._correo,
        "motivo": motivo,
        "fecha": time.strftime("%Y-%m-%d %H:%M:%S"),
        "estado": "Pendiente",
        "reporte": reporte
    })
    
    save_db() # Guardar base de datos
    return jsonify({"status": "success", "message": "Solicitud de retiro enviada correctamente al Coordinador."})

@app.route('/student/mark_notifications_read', methods=['POST'])
def student_mark_notifications_read():
    if 'usuario' not in session or session['rol'] != 'estudiante':
        return jsonify({"status": "error", "message": "No autorizado"})
    
    est = estudiantes[session['usuario']]
    for notif in est.notificaciones:
        notif['leido'] = True
    save_db()
    return jsonify({"status": "success"})


# -------------------------------------------------------------------------
# RUTA DOCENTE
# -------------------------------------------------------------------------
@app.route('/teacher')
def dashboard_docente():
    if 'usuario' not in session or session['rol'] != 'docente':
        return redirect(url_for('index'))
    
    doc = docentes.get(session['usuario'])
    if not doc:
        return redirect(url_for('logout'))
        
    carrera_id = getattr(doc, 'carrera_id', None)
    p_actual = periodos.get(carrera_id) or periodo_actual
    
    secciones_info = []
    for sec in doc.secciones:
        alumnos_sec = []
        for al in sec.estudiantes_inscritos:
            nota_obj = next((n for n in al.historial.lista_nota_materia if n.materia.id_materia == sec.materia.id_materia), None)
            
            alumnos_sec.append({
                "cedula": al.cedula,
                "nombre": al.obtener_nombre_completo(),
                "id_estudiante": al._id_estudiante,
                "parcial1": nota_obj.parcial1 if nota_obj else 0.0,
                "parcial2": nota_obj.parcial2 if nota_obj else 0.0,
                "asistencia": nota_obj.asistencia if nota_obj else 0,
                "nota_final": nota_obj.nota_final if nota_obj else 0.0,
                "estado_aprobacion": (nota_obj.esta_aprobado.value if nota_obj else "Sin Nota")
            })
            
        secciones_info.append({
            "id_seccion": sec.id_seccion,
            "materia": sec.materia.nombre_materia,
            "capacidad": sec.capacidad_estudiantil,
            "cupos_libres": sec.verificar_cupos_disponibles(),
            "inscritos_count": len(sec.estudiantes_inscritos),
            "estudiantes": alumnos_sec
        })

    mis_reportes = [r for r in reportes_generados if r.emisor == doc.obtener_nombre_completo()]
    
    # Calcular promedio
    promedio_doc = doc.ver_rendimiento()

    # Calcular horarios
    horarios_info = doc.ver_horario()

    return render_template(
        'dashboard_docente.html',
        docente=doc,
        secciones=secciones_info,
        promedio_rendimiento=promedio_doc,
        reportes=mis_reportes,
        estado_periodo=p_actual.estado_periodo,
        horarios=horarios_info
    )

@app.route('/teacher/student_info/<student_id>')
def teacher_student_info(student_id):
    if 'usuario' not in session or session['rol'] != 'docente':
        return jsonify({"status": "error", "message": "No autorizado"}), 403
        
    doc = docentes.get(session['usuario'])
    est = None
    if doc:
        est = next((e for sec in doc.secciones for e in sec.estudiantes_inscritos if e._id_estudiante == student_id), None)
    if not est:
        est = next((e for e in estudiantes.values() if e._id_estudiante == student_id), None)
    if not est:
        return jsonify({"status": "error", "message": "Estudiante no encontrado"}), 404
        
    notas_info = []
    for nota in est.historial.lista_nota_materia:
        notas_info.append({
            "materia": nota.materia.nombre_materia,
            "parcial1": nota.parcial1,
            "parcial2": nota.parcial2,
            "asistencia": nota.asistencia,
            "nota_final": nota.nota_final,
            "estado": nota.esta_aprobado.value
        })
        
    est_data = {
        "id_estudiante": est._id_estudiante,
        "nombres": est.nombres,
        "apellidos": est.apellidos,
        "correo": est._correo,
        "cedula": est.cedula,
        "nombre_periodo": est.nombre_periodo,
        "tipo_matricula": est._tipo_matricula,
        "notas": notas_info,
        "promedio": est.historial.promedio_general
    }
    
    return jsonify({"status": "success", "data": est_data})

@app.route('/teacher/save_grades', methods=['POST'])
def teacher_save_grades():
    if 'usuario' not in session or session['rol'] != 'docente':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    docente_obj = docentes.get(session['usuario'])
    if not docente_obj:
        return jsonify({"status": "error", "message": "Docente no encontrado."})

    carrera_id = getattr(docente_obj, 'carrera_id', None)
    p_actual = periodos.get(carrera_id) or periodo_actual
    
    if p_actual.estado_periodo != "En curso":
        return jsonify({"status": "error", "message": "No se pueden modificar calificaciones. El período académico no está en curso."})
        
    estudiante_id = request.form.get('estudiante_id')
    seccion_id = request.form.get('seccion_id')
    
    try:
        parcial1 = float(request.form.get('parcial1', 0.0))
        parcial2 = float(request.form.get('parcial2', 0.0))
        asistencia = int(request.form.get('asistencia', 0))
    except ValueError:
        return jsonify({"status": "error", "message": "Valores de calificaciones o asistencia inválidos."})

    if parcial1 < 0.0 or parcial1 > 10.0 or parcial2 < 0.0 or parcial2 > 10.0:
        return jsonify({"status": "error", "message": "Las calificaciones parciales deben estar entre 0.0 y 10.0."})
    if asistencia < 0 or asistencia > 100:
        return jsonify({"status": "error", "message": "La asistencia debe estar entre 0 y 100."})

    sec = secciones.get(seccion_id)
    est = None
    if sec:
        est = next((e for e in sec.estudiantes_inscritos if e._id_estudiante == estudiante_id), None)
    if not est:
        est = next((e for e in estudiantes.values() if e._id_estudiante == estudiante_id), None)

    if not est or not sec:
        return jsonify({"status": "error", "message": "Estudiante o Sección no encontrados."})

    nota_obj = next((n for n in est.historial.lista_nota_materia if n.materia.id_materia == sec.materia.id_materia), None)
    if not nota_obj:
        nota_obj = est.historial.crear_nota_materia(
            materia=sec.materia,
            periodo=p_actual,
            parcial1=0.0,
            parcial2=0.0,
            asistencia=0
        )
    
    docente_obj.colocar_calificacion(nota_obj, 1, parcial1)
    docente_obj.colocar_calificacion(nota_obj, 2, parcial2)
    docente_obj.tomar_asistencia(nota_obj, asistencia)

    save_db() # Guardar base de datos
    return jsonify({"status": "success", "message": f"Notas de {est.obtener_nombre_completo()} guardadas con éxito."})

@app.route('/teacher/generate_report', methods=['POST'])
def teacher_generate_report():
    if 'usuario' not in session or session['rol'] != 'docente':
        return jsonify({"status": "error", "message": "No autorizado"})
    
    doc = docentes[session['usuario']]
    tipo = request.form.get('tipo_reporte', 'Reporte de Notas')
    formato = request.form.get('formato', 'PDF')
    contenido = request.form.get('contenido', '').strip()
    
    if not contenido:
        return jsonify({"status": "error", "message": "El contenido del reporte no puede estar vacío."})
        
    reporte = doc.realizaReporte(tipo_de_reporte=tipo, formato_documento=formato, contenido=contenido)
    reportes_generados.append(reporte)
    
    save_db() # Guardar base de datos
    return jsonify({
        "status": "success",
        "message": "Reporte docente generado con éxito.",
        "report_content": reporte.imprimir_reporte()
    })


# -------------------------------------------------------------------------
# RUTA COORDINADOR
# -------------------------------------------------------------------------
@app.route('/coordinator')
def dashboard_coordinador():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return redirect(url_for('index'))
    
    coord = coordinadores[session['usuario']]
    carrera_id = coord.carrera.id_carrera if coord.carrera else None
    p_actual = periodos.get(carrera_id) or periodo_actual
    
    # Secciones
    secciones_info = []
    for sec in secciones.values():
        sec_carrera = getattr(sec, 'carrera_id', None)
        if sec_carrera is not None and sec_carrera != carrera_id:
            continue
        docentes_nombres = ", ".join(d.obtener_nombre_completo() for d in sec.docentes) if sec.docentes else "Sin Docente"
        secciones_info.append({
            "id_seccion": sec.id_seccion,
            "materia": sec.materia.nombre_materia,
            "capacidad": sec.capacidad_estudiantil,
            "inscritos": len(sec.estudiantes_inscritos),
            "docente": docentes_nombres,
            "aula": sec.aula_virtual._enlace_plataforma if sec.aula_virtual else "No asignada",
            "disponibilidad": "Disponible" if sec.disponibilidad else "Lleno",
            "tiene_horario": len(sec.lista_horarios) > 0,
            "horario_detalles": ", ".join(f"{h.turno} ({', '.join(h.dias)}: {h.hora_inicio}-{h.hora_fin})" for h in sec.lista_horarios) if sec.lista_horarios else "Sin Horario"
        })

    docentes_lista = [d for d in docentes.values() if getattr(d, 'carrera_id', None) is None or getattr(d, 'carrera_id', None) == carrera_id]
    materias_lista = [{"id": m.id_materia, "nombre": m.nombre_materia} for m in materias.values() if getattr(m, 'carrera_id', None) is None or getattr(m, 'carrera_id', None) == carrera_id]

    # Estadísticas
    estudiantes_activos = [e for e in estudiantes.values() if (getattr(e, 'carrera_id', None) is None or getattr(e, 'carrera_id', None) == carrera_id) and getattr(e, 'esta_activo', 1) == 1]
    total_estudiantes = len(estudiantes_activos)
    total_aprobados = sum(1 for e in estudiantes_activos if e.esta_aprobado == EstadoDeAprobacionNivelacion.APROBADO)
    total_reprobados = sum(1 for e in estudiantes_activos if e.esta_aprobado == EstadoDeAprobacionNivelacion.REPROBADO)
    total_pendientes = sum(1 for e in estudiantes_activos if e.esta_aprobado == EstadoDeAprobacionNivelacion.PENDIENTE)

    solicitudes_carrera = [s for s in solicitudes_retiro if s.get('id_carrera') is None or s.get('id_carrera') == carrera_id]

    return render_template(
        'dashboard_coordinador.html',
        coordinador=coord,
        periodo=p_actual,
        secciones=secciones_info,
        docentes=docentes_lista,
        materias=materias_lista,
        solicitudes=solicitudes_carrera,
        stats={
            "total": total_estudiantes,
            "aprobados": total_aprobados,
            "reprobados": total_reprobados,
            "pendientes": total_pendientes
        }
    )

@app.route('/coordinator/toggle_period', methods=['POST'])
def coordinator_toggle_period():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
    
    accion = request.form.get('accion')
    coord = coordinadores[session['usuario']]
    carrera_id = coord.carrera.id_carrera if coord.carrera else None
    p_actual = periodos.get(carrera_id)

    if not p_actual:
        return jsonify({"status": "error", "message": "Periodo no encontrado para esta carrera."})

    if accion == "iniciar":
        coord.abrir_periodo_matricula(p_actual)
        msg = "El periodo académico ha sido INICIADO. Los estudiantes pueden inscribirse y los docentes colocar notas."
    elif accion == "finalizar":
        career_students = [e for e in estudiantes.values() if getattr(e, 'carrera_id', None) == carrera_id]
        coord.cerrar_periodo_matricula(p_actual, career_students)
        msg = "El periodo académico ha sido FINALIZADO. Se han consolidado todas las notas finales y actas."
    else:
        return jsonify({"status": "error", "message": "Acción no válida"})

    save_db() # Guardar base de datos
    return jsonify({
        "status": "success", 
        "message": msg, 
        "estado_actual": p_actual.estado_periodo
    })

@app.route('/coordinator/approve_withdrawal', methods=['POST'])
def coordinator_approve_withdrawal():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    solicitud_id = int(request.form.get('solicitud_id', 0))
    accion = request.form.get('accion', '')
    
    sol = next((s for s in solicitudes_retiro if s['id'] == solicitud_id), None)
    if not sol:
        return jsonify({"status": "error", "message": "Solicitud no encontrada."})

    coord = coordinadores[session['usuario']]
    est = estudiantes.get(sol['correo'])
    
    aprobado = coord.aprobar_retiro(est, sol, accion)
    if aprobado:
        msg = f"Retiro aprobado. {sol['nombre']} ha sido desvinculado de todas las asignaturas."
    else:
        msg = f"Solicitud de retiro para {sol['nombre']} rechazada."

    save_db() # Guardar base de datos
    return jsonify({"status": "success", "message": msg})

@app.route('/coordinator/create_section', methods=['POST'])
def coordinator_create_section():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
    
    id_seccion = request.form.get('id_seccion', '').strip().upper()
    capacidad = request.form.get('capacidad')
    materia_id = request.form.get('materia_id')
    aula_link = request.form.get('aula_link', '').strip()
    aula_plataforma = request.form.get('aula_plataforma', 'TEAMS')
    
    if not id_seccion or not capacidad or not materia_id:
        return jsonify({"status": "error", "message": "Todos los campos son obligatorios para crear la sección."})
        
    try:
        capacidad = int(capacidad)
    except ValueError:
        return jsonify({"status": "error", "message": "La capacidad debe ser un número entero."})

    if id_seccion in secciones:
        return jsonify({"status": "error", "message": f"Ya existe una sección con el código {id_seccion}."})

    materia_obj = materias.get(materia_id)
    if not materia_obj:
        return jsonify({"status": "error", "message": "Materia no encontrada."})

    coord = coordinadores[session['usuario']]

    # Crear Aula Virtual si se proporciona enlace
    aula_obj = None
    if aula_link:
        if aula_plataforma.upper() == "ZOOM":
            servicio = ServicioZoom()
        else:
            servicio = ServicioTeams()
            
        is_examen = "examen" in id_seccion.lower()
        if is_examen:
            aula_obj = AulaExamen(40, servicio, aula_link)
        else:
            aula_obj = AulaClaseSincrona(40, servicio, aula_link)

    try:
        nueva_sec = (SeccionBuilder()
                    .con_id_seccion(id_seccion)
                    .con_capacidad_estudiantil(capacidad)
                    .con_materia(materia_obj)
                    .con_coordinador(coord)
                    .con_aula_virtual(aula_obj)
                    .con_entorno_asignado(aula_obj)
                    .build())
        nueva_sec.carrera_id = coord.carrera.id_carrera if coord.carrera else None
        
        secciones[id_seccion] = nueva_sec
        
        save_db() # Guardar base de datos
        return jsonify({"status": "success", "message": f"Sección {id_seccion} creada exitosamente para {materia_obj.nombre_materia}."})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error al crear sección con Builder: {e}"})

@app.route('/coordinator/update_section_aula', methods=['POST'])
def coordinator_update_section_aula():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    seccion_id = request.form.get('seccion_id')
    enlace = request.form.get('enlace', '').strip()
    plataforma = request.form.get('plataforma', 'TEAMS')
    
    if not seccion_id or seccion_id not in secciones:
        return jsonify({"status": "error", "message": "Sección no encontrada"})
        
    sec = secciones[seccion_id]
    
    if not enlace:
        sec.aula_virtual = None
        sec.entorno_asignado = None
        msg = f"Aula virtual desasignada de la sección {seccion_id}."
    else:
        if plataforma.upper() == "ZOOM":
            servicio = ServicioZoom()
        else:
            servicio = ServicioTeams()
            
        is_examen = "examen" in seccion_id.lower()
        if is_examen:
            aula = AulaExamen(40, servicio, enlace)
        else:
            aula = AulaClaseSincrona(40, servicio, enlace)
            
        sec.asignar_aula_virtual(aula)
        msg = f"Enlace de aula virtual actualizado para la sección {seccion_id}."
        
    save_db() # Guardar base de datos
    return jsonify({"status": "success", "message": msg})
@app.route('/coordinator/delete_all_students', methods=['POST'])
def coordinator_delete_all_students():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    correo_coord = session['usuario']
    coord = coordinadores.get(correo_coord)
    
    if not coord:
        return jsonify({"status": "error", "message": "Coordinador no encontrado"})
        
    auth_password = request.form.get('auth_password')
    if auth_password != coord.contrasenia:
        return jsonify({"status": "error", "message": "Contraseña incorrecta. Operación cancelada."})
        
    carrera_id = coord.carrera.id_carrera if coord.carrera else None
    if not carrera_id:
        return jsonify({"status": "error", "message": "No tienes una carrera asociada"})

    # Vaciar memoria de estudiantes de esta carrera
    for correo in list(estudiantes.keys()):
        if getattr(estudiantes[correo], 'carrera_id', None) == carrera_id:
            del estudiantes[correo]
    
    for sec in secciones.values():
        if getattr(sec, 'carrera_id', None) == carrera_id:
            sec.estudiantes_inscritos.clear()
            sec.disponibilidad = True
        
    # Eliminar archivos físicos de esta carrera
    import glob
    db_dir = os.path.join(os.path.dirname(__file__), 'database')
    students_dir = os.path.join(db_dir, carrera_id, 'students')
    if os.path.exists(students_dir):
        files = glob.glob(os.path.join(students_dir, '*.json'))
        for f in files:
            try:
                os.remove(f)
            except Exception as e:
                pass
                
    save_db()
    return jsonify({"status": "success", "message": "Todos los estudiantes de esta carrera y sus archivos han sido eliminados permanentemente."})

@app.route('/teacher/save_all_grades', methods=['POST'])
def teacher_save_all_grades():
    if 'usuario' not in session or session['rol'] != 'docente':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    docente_obj = docentes.get(session['usuario'])
    if not docente_obj:
        return jsonify({"status": "error", "message": "Docente no encontrado."})

    carrera_id = getattr(docente_obj, 'carrera_id', None)
    p_actual = periodos.get(carrera_id) or periodo_actual
    
    if p_actual.estado_periodo != "En curso":
        return jsonify({"status": "error", "message": "No se pueden modificar calificaciones. El período académico no está en curso."})
        
    data = request.json
    if not data or not isinstance(data, list):
        return jsonify({"status": "error", "message": "Datos inválidos. Se esperaba una lista de calificaciones."})

    for item in data:
        estudiante_id = item.get('estudiante_id')
        seccion_id = item.get('seccion_id')
        try:
            parcial1 = float(item.get('parcial1', 0.0))
            parcial2 = float(item.get('parcial2', 0.0))
            asistencia = int(item.get('asistencia', 0))
        except ValueError:
            return jsonify({"status": "error", "message": f"Valores de calificaciones o asistencia inválidos para {estudiante_id}."})
            
        if parcial1 < 0.0 or parcial1 > 10.0 or parcial2 < 0.0 or parcial2 > 10.0:
            return jsonify({"status": "error", "message": f"Las calificaciones parciales deben estar entre 0.0 y 10.0 para {estudiante_id}."})
        if asistencia < 0 or asistencia > 100:
            return jsonify({"status": "error", "message": f"La asistencia debe estar entre 0 y 100 para {estudiante_id}."})

        sec = secciones.get(seccion_id)
        est = None
        if sec:
            est = next((e for e in sec.estudiantes_inscritos if e._id_estudiante == estudiante_id), None)
        if not est:
            est = next((e for e in estudiantes.values() if e._id_estudiante == estudiante_id), None)

        if not est or not sec:
            continue

        nota_obj = next((n for n in est.historial.lista_nota_materia if n.materia.id_materia == sec.materia.id_materia), None)
        if not nota_obj:
            nota_obj = est.historial.crear_nota_materia(
                materia=sec.materia,
                periodo=p_actual,
                parcial1=0.0,
                parcial2=0.0,
                asistencia=0
            )

        nota_obj.parcial1 = parcial1
        nota_obj.parcial2 = parcial2
        nota_obj.asistencia = asistencia

    save_notas()
    return jsonify({"status": "success", "message": "Todas las calificaciones se han guardado exitosamente."})

@app.route('/coordinator/import_students', methods=['POST'])
def coordinator_import_students():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    seccion_id = request.form.get('seccion_id')
    if not seccion_id or seccion_id not in secciones:
        return jsonify({"status": "error", "message": "Sección no encontrada"})
        
    sec = secciones[seccion_id]
    
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "No se subió ningún archivo"})
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "Nombre de archivo no válido"})
        
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({"status": "error", "message": "El archivo debe ser un libro de Excel (.xlsx, .xls)"})
        
    import openpyxl
    from datetime import datetime
    import json
    import os
    
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        sheet = wb.active
        
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({"status": "error", "message": "El archivo de Excel está vacío o no contiene filas de datos"})
            
        headers = [str(h).strip().lower() for h in rows[0]]
        
        required_fields = ['cedula', 'nombres', 'apellidos', 'correo', 'contrasenia', 'id_estudiante', 'nombre_periodo', 'tipo_matricula']
        
        missing = [f for f in required_fields if f not in headers]
        if missing:
            return jsonify({"status": "error", "message": f"Faltan las siguientes columnas en el archivo: {', '.join(missing)}"})
            
        excel_students = []
        for row in rows[1:]:
            if all(val is None for val in row):
                continue
                
            student_data = {}
            for field in required_fields:
                col_idx = headers.index(field)
                val = row[col_idx]
                student_data[field] = str(val).strip() if val is not None else ""
                
            excel_students.append(student_data)
            
        # El número de estudiantes no debe exceder la capacidad de la sección
        if len(excel_students) > sec.capacidad_estudiantil:
            return jsonify({
                "status": "error", 
                "message": f"El número de estudiantes en el archivo ({len(excel_students)}) supera la capacidad de la sección ({sec.capacidad_estudiantil})."
            })
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_filename = f"estudiantes_{seccion_id}_{timestamp}.json"
        
        coord = coordinadores[session['usuario']]
        carrera_id = coord.carrera.id_carrera if coord.carrera else None
        p_actual = periodos.get(carrera_id) or periodo_actual

        imported_list = []
        for s_data in excel_students:
            correo = s_data['correo']
            
            est_obj = estudiantes.get(correo)
            if not est_obj:
                est_obj = Estudiante(
                    cedula=s_data['cedula'],
                    nombres=s_data['nombres'],
                    apellidos=s_data['apellidos'],
                    correo=correo,
                    contrasenia=s_data['contrasenia'],
                    id_estudiante=s_data['id_estudiante'],
                    nombre_periodo=s_data['nombre_periodo'],
                    tipo_matricula=s_data['tipo_matricula']
                )
                est_obj._archivo_origen = target_filename
                est_obj.carrera_id = carrera_id
                estudiantes[correo] = est_obj
            else:
                est_obj.cedula = s_data['cedula']
                est_obj.nombres = s_data['nombres']
                est_obj.apellidos = s_data['apellidos']
                est_obj.contrasenia = s_data['contrasenia']
                est_obj.nombre_periodo = s_data['nombre_periodo']
                est_obj._tipo_matricula = s_data['tipo_matricula']
                est_obj.carrera_id = carrera_id
                
            est_obj.esta_activo = 1
            est_obj.inscribir_seccion(sec)
                
            nota_obj = next((x for x in est_obj.historial.lista_nota_materia if x.materia.id_materia == sec.materia.id_materia), None)
            if not nota_obj:
                est_obj.historial.crear_nota_materia(
                    materia=sec.materia,
                    periodo=p_actual,
                    parcial1=0.0,
                    parcial2=0.0,
                    asistencia=0
                )
                
            imported_list.append(s_data)
            
        save_db()
            
        return jsonify({
            "status": "success", 
            "message": f"Se importaron con éxito {len(imported_list)} estudiantes a la sección {seccion_id} y se actualizó la base de datos."
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Error al procesar el archivo Excel: {e}"})

@app.route('/coordinator/import_teachers', methods=['POST'])
def coordinator_import_teachers():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "No se subió ningún archivo"})
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "Nombre de archivo no válido"})
        
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({"status": "error", "message": "El archivo debe ser un libro de Excel (.xlsx, .xls)"})
        
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({"status": "error", "message": "El archivo está vacío."})
            
        headers = [str(h).strip().lower() for h in rows[0]]
        required_fields = ['cedula', 'nombres', 'apellidos', 'correo', 'contrasenia', 'especialidad']
        
        missing = [f for f in required_fields if f not in headers]
        if missing:
            return jsonify({"status": "error", "message": f"Faltan columnas: {', '.join(missing)}"})
            
        imported_count = 0
        for row in rows[1:]:
            if all(val is None for val in row):
                continue
            
            d_data = {}
            for field in required_fields:
                col_idx = headers.index(field)
                val = row[col_idx]
                d_data[field] = str(val).strip() if val is not None else ""
                
            coord = coordinadores[session['usuario']]
            carrera_id = coord.carrera.id_carrera if coord.carrera else None
            correo = d_data['correo']
            
            if correo not in docentes:
                doc_obj = Docente(
                    cedula=d_data['cedula'],
                    nombres=d_data['nombres'],
                    apellidos=d_data['apellidos'],
                    correo=correo,
                    contrasenia=d_data['contrasenia'],
                    especialidad=d_data['especialidad']
                )
                doc_obj.carrera_id = carrera_id
                docentes[correo] = doc_obj
            else:
                doc_obj = docentes[correo]
                doc_obj.cedula = d_data['cedula']
                doc_obj.nombres = d_data['nombres']
                doc_obj.apellidos = d_data['apellidos']
                doc_obj.contrasenia = d_data['contrasenia']
                doc_obj.especialidad = d_data['especialidad']
                doc_obj.carrera_id = carrera_id
                
            imported_count += 1
            
        save_db()
        return jsonify({"status": "success", "message": f"Se importaron {imported_count} docentes exitosamente."})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Error procesando Excel: {e}"})

@app.route('/coordinator/import_curriculum', methods=['POST'])
def coordinator_import_curriculum():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    coord = coordinadores[session['usuario']]
    carrera = coord.carrera
    if not carrera:
        return jsonify({"status": "error", "message": "No tienes una carrera asociada"})
        
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "No se subió ningún archivo"})
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "Nombre de archivo no válido"})
        
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({"status": "error", "message": "El archivo debe ser un libro de Excel (.xlsx, .xls)"})
        
    import openpyxl
    import json
    import os
    
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        sheet = wb.active
        
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({"status": "error", "message": "El archivo de Excel está vacío o no contiene filas de datos"})
            
        headers = [str(h).strip().lower() for h in rows[0]]
        required_fields = ['id_materia', 'nombre_materia', 'nota_minima', 'asistencia_minima']
        
        missing = [f for f in required_fields if f not in headers]
        if missing:
            return jsonify({"status": "error", "message": f"Faltan las siguientes columnas: {', '.join(missing)}"})
            
        excel_materias = []
        for row in rows[1:]:
            if all(val is None for val in row):
                continue
            mat_data = {}
            for field in required_fields:
                col_idx = headers.index(field)
                val = row[col_idx]
                mat_data[field] = val
            excel_materias.append(mat_data)
            
        if len(excel_materias) != 5:
            return jsonify({"status": "error", "message": f"La malla curricular debe contener exactamente 5 materias, se encontraron {len(excel_materias)}."})
            
        # Actualizar global materias (solo de esta carrera)
        for k in list(materias.keys()):
            if getattr(materias[k], 'carrera_id', None) == carrera.id_carrera:
                del materias[k]

        for m_data in excel_materias:
            try:
                nota_minima = float(m_data["nota_minima"])
                asistencia_minima = int(m_data["asistencia_minima"])
            except ValueError:
                return jsonify({"status": "error", "message": "Error de formato: nota_minima debe ser número y asistencia_minima entero."})
                
            m = Materia(
                id_materia=str(m_data["id_materia"]).strip(),
                nombre_materia=str(m_data["nombre_materia"]).strip(),
                nota_minima=nota_minima,
                asistencia_minima=asistencia_minima
            )
            m.carrera_id = carrera.id_carrera
            materias[m.id_materia] = m
            
        # Guardar JSON en la carpeta de la carrera
        career_dir = os.path.join(os.path.dirname(__file__), 'database', carrera.id_carrera)
        mallas_dir = os.path.join(career_dir, 'mallas_curriculares')
        os.makedirs(mallas_dir, exist_ok=True)
        json_filename = f"{carrera.nombre_carrera}.json"
        json_path = os.path.join(mallas_dir, json_filename)
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(excel_materias, f, indent=2, ensure_ascii=False)
            
        save_db()
        
        return jsonify({
            "status": "success",
            "message": f"Se importaron con éxito 5 materias y se creó el archivo {json_filename}."
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Error al procesar el archivo Excel: {str(e)}"})

@app.route('/coordinator/assign_schedule', methods=['POST'])
def coordinator_assign_schedule():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    seccion_id = request.form.get('seccion_id')
    hora_inicio = request.form.get('hora_inicio')
    hora_fin = request.form.get('hora_fin')
    turno = request.form.get('turno', 'Asignado')
    modalidad = request.form.get('modalidad', 'PRESENCIAL').upper()
    dias = request.form.getlist('dias')
    
    if not seccion_id or not hora_inicio or not hora_fin:
        return jsonify({"status": "error", "message": "Todos los campos de la hora son obligatorios."})
        
    if not dias:
        return jsonify({"status": "error", "message": "Debe seleccionar al menos un día de la semana."})
        
    sec = secciones.get(seccion_id)
    if not sec:
        return jsonify({"status": "error", "message": "Sección no encontrada."})
        
    if hora_inicio >= hora_fin:
        return jsonify({"status": "error", "message": "La hora de fin debe ser posterior a la hora de inicio."})

    coord = coordinadores[session['usuario']]
    nuevo_horario = Horario(turno, hora_inicio, hora_fin, modalidad, dias)
    
    try:
        coord.asignar_horario_a_seccion(sec, nuevo_horario, secciones.values())
    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)})
    
    save_db() # Guardar base de datos
    return jsonify({"status": "success", "message": f"Horario asignado con éxito a la sección {sec.id_seccion}."})

@app.route('/coordinator/assign_teacher', methods=['POST'])
def coordinator_assign_teacher():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    seccion_id = request.form.get('seccion_id')
    docente_correo = request.form.get('docente_correo')
    
    sec = secciones.get(seccion_id)
    doc = docentes.get(docente_correo)
    
    if not sec or not doc:
        return jsonify({"status": "error", "message": "Sección o Docente no encontrados."})
        
    coord = coordinadores[session['usuario']]
    try:
        coord.asignar_docente_a_seccion(doc, sec)
    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)})
    
    save_db() # Guardar base de datos
    return jsonify({"status": "success", "message": f"Docente {doc.obtener_nombre_completo()} asignado con éxito a la sección {sec.id_seccion}."})

@app.route('/coordinator/generate_career_report', methods=['POST'])
def coordinator_generate_career_report():
    if 'usuario' not in session or session['rol'] != 'coordinador':
        return jsonify({"status": "error", "message": "No autorizado"})
        
    formato = request.form.get('formato', 'PDF')
    
    carrera_software.estudiantes_inscritos = sum(1 for e in estudiantes.values() if getattr(e, 'esta_activo', 1) == 1)
    reporte = carrera_software.mostrar_datos_carrera(formato_documento=formato)
    reportes_generados.append(reporte)
    
    save_db() # Guardar base de datos
    return jsonify({
        "status": "success",
        "message": "Reporte de carrera generado con éxito.",
        "report_content": reporte.imprimir_reporte()
    })

# =========================================================================
# LANZADOR DE LA APLICACIÓN
# =========================================================================

# Cargar datos al iniciar
try:
    load_db()
    print("Base de datos cargada correctamente.")
except Exception as e:
    print(f"Error cargando la base de datos: {e}")

if __name__ == '__main__':
    port = 5000
    
    flask_thread = threading.Thread(
        target=lambda: app.run(host='127.0.0.1', port=port, debug=False, use_reloader=False),
        daemon=True
    )
    flask_thread.start()
    
    print("--------------------------------------------------")
    print(f" Servidor local activo en: http://127.0.0.1:{port}")
    print(" Intentando iniciar ventana WebView nativa...")
    print("--------------------------------------------------")
    
    time.sleep(1.2)
    
    try:
        import webview
        webview.create_window(
            title='Sistema de Gestión de Nivelación - ULEAM', 
            url=f'http://127.0.0.1:{port}', 
            width=1200, 
            height=820,
            resizable=True,
            min_size=(1000, 600)
        )
        webview.start()
    except Exception as e:
        print(f"\n[AVISO] No se pudo lanzar pywebview automáticamente ({e}).")
        print("Abriendo en el navegador predeterminado...")
        webbrowser.open(f'http://127.0.0.1:{port}')
        
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\nServidor detenido por el usuario.")
